import logging
import os
import openpyxl
import xml.etree.ElementTree as ET
import csv
from app.tabletki_data_service.app.services.data_converter import process_data_converter  # Импорт data_converter
from app.database import get_async_db
from app.notification_service import send_notification  # Импортируем функцию для отправки уведомлений

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

NOTIFICATION_FILE = "notifications.txt"  # Файл для уведомлений
def validate_consistency(data, file_type, single_store, store_serial,enterprise_code):
    """
    Проверяет согласованность данных для branch и store_serial.
    :param data: Список строк из файла.
    :param file_type: Тип файла (catalog или stock).
    :param single_store: Используется ли один магазин (True/False).
    :param store_serial: Уникальный идентификатор филиала.
    """
    if file_type == "catalog":  # Проверка для каталога
        seen_codes = set()
        for row in data:
            code = row.get("code", "").strip()
            if not code:
                continue  # Игнорируем строки с пустым или пробельным значением code
            name = row.get("name", "").strip()
            producer = row.get("producer", "").strip()
            # Проверка на наличие одинаковых значений в поле "code"
            if code in seen_codes:
                message = f"Ошибка: Найдено дублирование значения code '{code}' в файле."
                logging.warning(message)
                send_notification(f"Ошибка в каталоге для предприятия {enterprise_code}", message)
            else:
                seen_codes.add(code)

            # Если code есть, а name или producer пустые
            if code and not name:
                message = f"Ошибка: Для кода '{code}' отсутствует имя (name)."
                logging.warning(message)
                send_notification(f"Ошибка в каталоге для предприятия {enterprise_code}", message)
            
            if code and not producer:
                message = f"Ошибка: Для кода '{code}' отсутствует производитель (producer)."
                logging.warning(message)
                send_notification(f"Ошибка в каталоге для предприятия {enterprise_code}", message)

            # Продолжаем проверку других строк
            continue
    if file_type == "stock":  # Проверка актуальна только для stock
        for row in data:
            code = row.get("code", "").strip()
            if code:  # Если значение поля code есть
                # Проверяем наличие других обязательных полей, если есть значение в code
                required_fields = ["price", "qty", "pricereserve"]
                missing_fields = [field for field in required_fields if not row.get(field, "").strip()]
                if missing_fields:
                    message = f"Ошибка: Для строки с кодом '{code}' отсутствуют обязательные поля: {', '.join(missing_fields)}."
                    logging.error(message)
                    send_notification(f"Ошибка в файле stock для предприятия {enterprise_code}", message)
                    return  # Прерываем обработку файла, если обязательные поля отсутствуют
            try:
                price = float(row.get("price", 0))  # Получаем Price
                price_reserve = float(row.get("pricereserve", 0))  # Получаем PriceReserve

                if price_reserve > price:
                    message = f"Ошибка: PriceReserve ({price_reserve}) не может быть больше Price ({price})."
                    send_notification(f"Ошибка в строке {row} для предприятия-{enterprise_code}",message )
                    logging.warning(f"Найдено несоответствие: {message}")
                    
                    continue # Пропускаем этот товар, но продолжаем обработку файла
            except ValueError as e:
                logging.error(f"Ошибка в строке: {row}. Детали: {e}")
               
                continue  # Пропускаем строку с ошибкой, но продолжаем обработку файла
   
    if file_type != "stock" or single_store:
        return  # Проверка не нужна для catalog или если single_store=True

    branch_to_serial = {}
    for row in data:
        branch = row.get("branch", "").strip()
        serial = row.get("store_serial", "").strip()

        # Проверяем соответствие branch и store_serial
        if branch and serial:
            if serial in branch_to_serial and branch_to_serial[serial] != branch:
                raise ValueError(f"Конфликт данных: store_serial '{serial}' уже связан с другим branch '{branch_to_serial[serial]}'.")
            branch_to_serial[serial] = branch

    logging.info("Проверка согласованности данных завершена.")

async def validate_data(enterprise_code, file_path, file_type, single_store, store_serial):
    """
    Основная функция проверки данных.
    :param enterprise_code: Код предприятия
    :param file_path: Путь к файлу
    :param file_type: Тип файла (catalog или stock)
    :param single_store: Указание на использование single_store
    :param store_serial: Уникальный идентификатор филиала
    """
    try:
        logging.info(f"Параметры перед проверкой: file_path={file_path}, file_type={file_type}, "
                     f"single_store={single_store}, store_serial={store_serial}")

        # Проверка наличия store_serial для single_store режима
        if single_store and not store_serial:
            raise ValueError("store_serial обязателен для single_store режима.")

        # Проверка обязательности Branch для multi-store режима
        if not single_store and file_type == "stock":
            data = read_file_data(file_path, file_type)
            for row in data:

                branch = row.get("branch", "").strip()
                if not branch:
                    message = f"Ошибка: {row}Для multi-store режима в файле stock обязательно должно быть указано поле 'branch'."
                    logging.warning(f"Найдено несоответствие: {message}")
                    send_notification(f"Ошибка для предприятия {enterprise_code}",message )
                    continue  # Пропускаем этот товар, но продолжаем обработку файла

        # Проверка формата файла
        if file_path.endswith(".xlsx") or file_path.endswith(".xls"):
            validate_excel(file_path, file_type, single_store, store_serial,enterprise_code)
        elif file_path.endswith(".xml"):
            validate_xml(file_path, file_type, single_store, store_serial,enterprise_code)
        elif file_path.endswith(".csv"):
            validate_csv(file_path, file_type, single_store, store_serial,enterprise_code)
        else:
            raise ValueError(f"Неизвестный формат файла: {file_path}")

        success_message = f"Файл {file_path} успешно прошел проверку для предприятия {enterprise_code}."
        logging.info(success_message)
        

        # Интеграция с data_converter
        logging.info(f"Передача файла {file_path} в data_converter для обработки.")
        
        # Создаем сессию базы данных
        async with get_async_db() as db_session:
            logging.info(f"Тип и значение enterprise_code перед вызовом process_data_converter: {type(enterprise_code)} - {enterprise_code}")
            await process_data_converter(
                enterprise_code=enterprise_code,
                file_path=file_path,
                file_type=file_type,
                store_serial=store_serial,
                single_store=single_store,
                db_session=db_session  # Передача сессии базы данных
            )
        
        logging.info(f"Файл {file_path} успешно обработан data_converter для предприятия {enterprise_code}.")
        return True

    except Exception as e:
        error_message = f"Ошибка валидации файла {file_path} для предприятия {enterprise_code}: {str(e)}"
        logging.error(error_message)
        send_notification(f"Внимание",error_message)
        
        return False

def read_file_data(file_path, file_type):
    """
    Считывает данные из файла в общий формат (универсальный парсер).
    :param file_path: Путь к файлу.
    :param file_type: Тип файла (catalog или stock).
    :return: Список строк с данными.
    """
    if file_path.endswith(".xlsx") or file_path.endswith(".xls"):
        return read_excel_data(file_path)
    elif file_path.endswith(".xml"):
        return read_xml_data(file_path)
    elif file_path.endswith(".csv"):
        return read_csv_data(file_path)
    else:
        raise ValueError(f"Неизвестный формат файла: {file_path}")

def read_excel_data(file_path):
    """
    Считывает данные из Excel в общий формат.
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in sheet[1]]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[col_idx]: str(cell).strip() if cell is not None else "" for col_idx, cell in enumerate(row)}
        # Пропускаем строки, в которых нет значимых данных (все значения пустые)
        if all(value == "" for value in row_data.values()):
            continue  # Игнорируем эту строку, если все значения пустые
        data.append(row_data)
    return data

def read_xml_data(file_path):
    """
    Считывает данные из XML в общий формат с нормализацией ключей.
    """
    tree = ET.parse(file_path)
    root = tree.getroot()

    # Сбор данных из тегов <Item> (или другого актуального тега)
    data = []
    for element in root.findall("Item"):  # Убедитесь, что тег "Item" соответствует структуре XML
        row = {}
        for child in element:
            key = child.tag.strip()  # Оригинальный регистр ключа
            value = child.text.strip() if child.text else None
            row[key] = value
        data.append(row)

    # Преобразуем Price и PriceReserve в числа
    for row in data:
        if "Price" in row:
            row["Price"] = float(row["Price"]) if row["Price"] else 0.0
        if "PriceReserve" in row:
            row["PriceReserve"] = float(row["PriceReserve"]) if row["PriceReserve"] else 0.0

    logging.info(f"Данные, извлеченные из XML: {data}")
    return data

def read_csv_data(file_path):
    """
    Считывает данные из CSV в общий формат.
    """
    with open(file_path, mode="r", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        headers = [header.strip().lower() for header in reader.fieldnames]  # Приводим заголовки к нижнему регистру
        data = []
        for row in reader:
            cleaned_row = {k.strip().lower(): v.strip() for k, v in row.items() if v and v.strip()}  # Приводим ключи и значения
            if cleaned_row:
                data.append(cleaned_row)

        logging.info(f"CSV файл {file_path} содержит {len(data)} строк.")

        return data

def validate_excel(file_path: str, file_type: str, single_store: bool, store_serial: str,enterprise_code):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Определение обязательных полей
        required_fields = {
            "catalog": ["code", "name", "producer"],  # Приводим обязательные поля к нижнему регистру
            "stock": ["code", "price", "qty", "pricereserve"]
        }[file_type]

        if file_type == "stock" and not single_store:
            required_fields.append("branch")

        # Чтение заголовков из первой строки и приведение их к нижнему регистру
        headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in sheet[1]]
        

        # Проверка отсутствующих полей
        missing_fields = [field for field in required_fields if field not in headers]
        if missing_fields:
            raise ValueError(f"Отсутствуют обязательные поля: {', '.join(missing_fields)}")

        

        # Чтение данных с учетом заголовков
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {headers[col_idx]: str(cell).strip() if cell is not None else "" for col_idx, cell in enumerate(row)}
            data.append(row_data)

        # Проверка консистентности данных
        validate_consistency(data, file_type, single_store, store_serial,enterprise_code)

    except Exception as e:
        error_message = f"Ошибка валидации Excel файла {file_path}: {str(e)}для предприятия {enterprise_code}"
        logging.error(error_message)
        send_notification(f"Внимание",error_message)
        raise
def validate_xml(file_path: str, file_type: str, single_store: bool, store_serial: str,enterprise_code):
    """
    Проверяет данные в файле XML.
    """
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()

        # Определение обязательных полей
        required_fields = {
            "catalog": ["code", "name", "producer"],
            "stock": ["code", "price", "qty", "pricereserve"]
        }[file_type]

        if file_type == "stock" and not single_store:
            required_fields.append("branch")

        # Сбор данных
        data = [
            {child.tag.lower(): (child.text.strip() if child.text else "") for child in element}
            for element in root.findall("row")
        ]

        # Преобразование значений price и pricereserve в float
        for row in data:
            if "price" in row:
                row["price"] = float(row["price"]) if row["price"] else 0.0
            if "pricereserve" in row:
                row["pricereserve"] = float(row["pricereserve"]) if row["pricereserve"] else 0.0

        # Логгирование количества строк
        logging.info(f"XML файл {file_path} содержит {len(data)} строк: {data}")

        # Проверка отсутствующих полей
        for row in data:
            missing_fields = [field for field in required_fields if field not in row]
            if missing_fields:
                raise ValueError(f"Отсутствуют обязательные поля: {', '.join(missing_fields)}")

        logging.info(f"Файл {file_path} успешно прошел проверку заголовков.")

        # Проверка консистентности данных
        validate_consistency(data, file_type, single_store, store_serial,enterprise_code)

    except Exception as e:
        error_message = f"Ошибка валидации XML файла {file_path}: {str(e)}для предприятия {enterprise_code}"
        logging.error(error_message)
        send_notification(f"Внимание",error_message)
        raise
def validate_csv(file_path: str, file_type: str, single_store: bool, store_serial: str,enterprise_code):
    """
    Проверяет данные в файле CSV.
    """
    try:
        with open(file_path, mode="r", encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            headers = reader.fieldnames

            # Определение обязательных полей
            required_fields = {
                "catalog": ["code", "name", "producer"],
                "stock": ["code", "price", "Qty", "PriceReserve"]
            }[file_type]

            if file_type == "stock" and not single_store:
                required_fields.append("branch")

            # Проверка отсутствующих полей
            missing_fields = [field for field in required_fields if field not in headers]
            if missing_fields:
                raise ValueError(f"Отсутствуют обязательные поля: {', '.join(missing_fields)}")

            logging.info(f"Файл {file_path} успешно прошел проверку заголовков.")

            # Чтение данных
            data = []
            for row in reader:
                # Убираем пробелы и проверяем на пустоту
                cleaned_row = {k.strip(): v.strip() for k, v in row.items() if v and v.strip()}
                if cleaned_row:
                    data.append(cleaned_row)

            # Логгирование количества строк
            logging.info(f"CSV файл {file_path} содержит {len(data)} строк.")

            # Проверка консистентности данных
            validate_consistency(data, file_type, single_store, store_serial,enterprise_code)

    except Exception as e:
        error_message = f"Ошибка валидации CSV файла {file_path}: {str(e)}для предприятия {enterprise_code}"
        logging.error(error_message)
        send_notification(f"Внимание",error_message)
        
        raise  