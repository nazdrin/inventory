from __future__ import annotations

import asyncio
import io
import json
import logging
import os
from typing import Optional, List, Dict, Literal, Any

import httpx
import xml.etree.ElementTree as ET
from sqlalchemy import text

from app.database import get_async_db
from app.services.notification_service import send_notification

# Google Drive
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload

# Excel
from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# === ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ XML/ФИДА D5 ===


def _get_text(el: ET.Element, candidates: List[str]) -> Optional[str]:
    """Возвращает текст первого дочернего тега из списка кандидатов."""
    for tag in candidates:
        child = el.find(tag)
        if child is not None and child.text and child.text.strip():
            return child.text.strip()
    return None


def _to_float(val: Optional[str]) -> float:
    """Мягко преобразует строку в float (учитывая пробелы/запятые)."""
    if not val:
        return 0.0
    s = str(val).strip()
    s = s.replace(" ", "").replace("\u00A0", "")
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


async def _get_feed_url_by_code(code: str = "D5") -> Optional[str]:
    """Достаёт feed_url из dropship_enterprises по значению поля code."""
    async with get_async_db() as session:
        res = await session.execute(
            text(
                "SELECT feed_url "
                "FROM dropship_enterprises "
                "WHERE code = :code "
                "LIMIT 1"
            ),
            {"code": code},
        )
        return res.scalar_one_or_none()


async def _get_retail_markup_by_code(code: str) -> float:
    """Возвращает retail_markup (в процентах) из dropship_enterprises по значению поля code.

    Если значение не найдено или невалидно, возвращает 0.0.
    """
    async with get_async_db() as session:
        res = await session.execute(
            text(
                "SELECT retail_markup "
                "FROM dropship_enterprises "
                "WHERE code = :code "
                "LIMIT 1"
            ),
            {"code": code},
        )
        value = res.scalar_one_or_none()

    if value is None:
        return 0.0

    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


async def _get_profit_percent_by_code(code: str) -> float:
    """Возвращает profit_percent из dropship_enterprises по code.

    В БД значение хранится как проценты (например, 10), возвращаем долю (0.1).
    Если не найдено или невалидно — 0.0.
    """
    async with get_async_db() as session:
        res = await session.execute(
            text(
                "SELECT profit_percent "
                "FROM dropship_enterprises "
                "WHERE code = :code "
                "LIMIT 1"
            ),
            {"code": code},
        )
        value = res.scalar_one_or_none()

    if value is None:
        return 0.0

    try:
        val = float(value)
    except (TypeError, ValueError):
        return 0.0

    # 10 -> 0.1
    if val > 1:
        val = val / 100.0

    # нормализация
    if val < 0:
        val = 0.0
    if val > 1:
        val = 1.0

    return val


async def _load_feed_root(*, code: str, timeout: int) -> Optional[ET.Element]:
    """
    Получение XML фида:
    1) берём feed_url из БД по code
    2) скачиваем XML
    3) возвращаем корень ElementTree
    """
    feed_url = await _get_feed_url_by_code(code)
    if not feed_url:
        msg = f"Не найден feed_url в dropship_enterprises для code='{code}'"
        logger.error(msg)
        send_notification(msg, "Разработчик")
        return None

    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        async with httpx.AsyncClient(headers=headers, timeout=timeout) as client:
            resp = await client.get(feed_url)
            resp.raise_for_status()
            xml_text = resp.text
    except Exception as e:
        msg = f"Ошибка загрузки фида {feed_url}: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        return None

    try:
        return ET.fromstring(xml_text)
    except Exception as e:
        msg = f"Ошибка парсинга XML из {feed_url}: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        return None


def _collect_offer_nodes(root: ET.Element) -> List[ET.Element]:
    """Собираем узлы <offer> из фида."""
    offers = root.findall(".//offer")
    if not offers:
        # Фолбек: возьмём элементы, похожие на товары
        offers = [el for el in root.iter() if el.tag.lower() in ("offer", "item")]
    return offers


# === GOOGLE DRIVE / EXCEL ДЛЯ КАТАЛОГА D5 ===


async def _connect_to_google_drive():
    """
    Создает клиент Drive API через сервисный аккаунт.
    Использует GOOGLE_DRIVE_CREDENTIALS_PATH.
    """
    creds_path = os.getenv("GOOGLE_DRIVE_CREDENTIALS_PATH")
    if not creds_path or not os.path.exists(creds_path):
        msg = f"Неверный путь к учетным данным Google Drive: {creds_path}"
        logger.error(msg)
        send_notification(msg, "Разработчик")
        raise FileNotFoundError(msg)

    credentials = service_account.Credentials.from_service_account_file(
        creds_path,
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    service = build("drive", "v3", credentials=credentials)
    logger.info("Подключено к Google Drive")
    return service


async def _fetch_single_file_metadata(drive_service, folder_id: str) -> Dict[str, Any]:
    """
    Возвращает metadata ОДНОГО файла (id, name) из папки.
    По условию задачи в папке всегда один файл.
    """
    try:
        results = (
            drive_service.files()
            .list(
                q=f"'{folder_id}' in parents and trashed=false",
                fields="files(id, name)",
                pageSize=10,
            )
            .execute()
        )
        files = results.get("files", []) or []
        if not files:
            raise FileNotFoundError(f"В папке {folder_id} нет файлов")
        return files[0]
    except HttpError as e:
        msg = f"HTTP ошибка при получении файлов из папки {folder_id}: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        raise


async def _download_file_bytes(drive_service, file_id: str) -> bytes:
    """
    Загружает файл с Google Drive в bytes.
    """
    try:
        request = drive_service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)
        return fh.read()
    except HttpError as e:
        msg = f"HTTP ошибка при загрузке файла {file_id}: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        raise


async def _get_gdrive_folder_id_by_code(code: str) -> str:
    """
    Берёт ID папки с Excel для каталога из таблицы dropship_enterprises.gdrive_folder.
    """
    async with get_async_db() as session:
        res = await session.execute(
            text(
                "SELECT gdrive_folder "
                "FROM dropship_enterprises "
                "WHERE code = :code "
                "LIMIT 1"
            ),
            {"code": code},
        )
        folder_id = res.scalar_one_or_none()

    if not folder_id:
        msg = f"Не найден gdrive_folder в dropship_enterprises для code='{code}'"
        logger.error(msg)
        send_notification(msg, "Разработчик")
        raise RuntimeError(msg)

    return str(folder_id)


def _parse_catalog_excel_bytes(data: bytes) -> List[Dict[str, str]]:
    """
    Парсит Excel-файл из bytes и возвращает список:
    [{"id": ..., "name": ..., "barcode": ...}, ...]
    Маппинг:
      Артикул      -> id
      Номенклатура -> name
      Штрихкод     -> barcode
    """
    wb = load_workbook(io.BytesIO(data), data_only=True)
    sheet = wb.active

    # Первая строка — заголовки
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {
        str(value).strip(): idx
        for idx, value in enumerate(header_row)
        if value is not None
    }

    required_cols = ["Артикул", "Номенклатура", "Штрихкод"]
    for col in required_cols:
        if col not in headers:
            raise ValueError(f"В Excel не найден обязательный столбец '{col}'")

    idx_artikul = headers["Артикул"]
    idx_name = headers["Номенклатура"]
    idx_barcode = headers["Штрихкод"]

    items: List[Dict[str, str]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        artikul = row[idx_artikul] if idx_artikul < len(row) else None
        name = row[idx_name] if idx_name < len(row) else None
        barcode = row[idx_barcode] if idx_barcode < len(row) else None

        if not artikul or not name:
            # Без Артикул/Номенклатуры строку пропускаем
            continue

        items.append(
            {
                "id": str(artikul).strip(),
                "name": str(name).strip(),
                "barcode": (str(barcode).strip() if barcode else ""),
            }
        )

    logger.info("Каталог (Excel D5): собрано позиций: %d", len(items))
    return items


# === ОСНОВНЫЕ ФУНКЦИИ D5 ===


async def parse_feed_catalog_to_json(
    *,
    code: str = "D5",
    timeout: int = 30,  # не используется, но оставлен для совместимости
) -> str:
    """
    Каталог для D5:
    берём ОДИН Excel файл из папки на Google Drive.
    ID папки берётся из dropship_enterprises.gdrive_folder по code.
    Читаем столбцы:
      Артикул      -> id
      Номенклатура -> name
      Штрихкод     -> barcode
    Возвращаем JSON со списком:
    [
      {"id": "<Артикул>", "name": "<Номенклатура>", "barcode": "<Штрихкод>"},
      ...
    ]
    """
    try:
        drive_service = await _connect_to_google_drive()
        folder_id = await _get_gdrive_folder_id_by_code(code)
        file_meta = await _fetch_single_file_metadata(drive_service, folder_id)
        file_id = file_meta["id"]
        file_name = file_meta.get("name")
        logger.info("Найден файл каталога для %s: %s (%s)", code, file_name, file_id)

        file_bytes = await _download_file_bytes(drive_service, file_id)
        items = _parse_catalog_excel_bytes(file_bytes)
    except Exception as e:
        msg = f"Ошибка обработки каталога {code} из Google Drive: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        return "[]"

    return json.dumps(items, ensure_ascii=False, indent=2)


async def parse_feed_stock_to_json(
    *,
    code: str = "D5",
    timeout: int = 30,
) -> str:
    """
    Сток для D5: возвращает JSON со списком
    [
      {"code_sup": "<vendorCode>", "qty": <0|1>, "price_retail": <float>, "price_opt": 0},
      ...
    ]

    Маппинг:
      - vendorCode          -> code_sup
      - available="true"    -> qty=1, иначе 0
      - price               -> base_price
      - price_opt           -> base_price * (1 - profit_percent)
      - price_retail        -> price_opt * (1 + retail_markup/100)

    Пример входного XML:
      <offer id="39" group_id="4712433" available="true">
        <url>...</url>
        <oldprice>1995</oldprice>
        <price>1795</price>
        ...
        <vendorCode>OPN-02956</vendorCode>
        <vendor>Optimum Nutrition</vendor>
        <name><![CDATA[ ON Gold Standard 100% Whey Protein 900 грам (EU), Банан ]]></name>
      </offer>
    """
    root = await _load_feed_root(code=code, timeout=timeout)
    if root is None:
        return "[]"

    # Наценка для розницы из dropship_enterprises.retail_markup (в процентах)
    try:
        retail_markup = await _get_retail_markup_by_code(code)
    except Exception as e:
        msg = f"Ошибка получения retail_markup для code='{code}': {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        retail_markup = 0.0

    # Профит (процент) из dropship_enterprises.profit_percent (например 10 -> 0.1)
    try:
        profit = await _get_profit_percent_by_code(code)
    except Exception as e:
        msg = f"Ошибка получения profit_percent для code='{code}': {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        profit = 0.0

    # Коэффициент: 1 + retail_markup/100, например при 10% будет 1.10
    retail_coef = 1.0 + (retail_markup / 100.0 if retail_markup else 0.0)

    rows: List[Dict[str, Any]] = []

    for offer in _collect_offer_nodes(root):
        # vendorCode — дочерний тег или атрибут (на всякий случай поддержим оба)
        vendor_code = (
            _get_text(offer, ["vendorCode"]) or offer.get("vendorCode")
        )
        if not vendor_code:
            # Без vendorCode — пропускаем
            continue

        # available="true" -> qty = 1, иначе 0
        available_raw = (offer.get("available") or "").strip().lower()
        qty = 1 if available_raw == "true" else 0

        # Игнорируем позиции с нулевым или отрицательным остатком
        if qty <= 0:
            continue

        price_raw = _get_text(offer, ["price"])
        base_price = _to_float(price_raw)

        # Оптовая цена: price_opt = base_price * (1 - profit)
        # profit хранится в БД как проценты (например 10), выше преобразовано в долю (0.1)
        price_opt = base_price / (1.0 + profit) if base_price > 0 else 0.0
        if price_opt < 0:
            price_opt = 0.0

        # Розница: adjusted_price = price_opt * (1 + retail_markup/100)
        adjusted_price = price_opt * retail_coef if price_opt > 0 else 0.0

        # В текущем формате стока используем целые грн
        price_opt_int = int(price_opt)
        price_retail_int = int(adjusted_price)

        rows.append(
            {
                "code_sup": str(vendor_code).strip(),
                "qty": qty,
                "price_retail": price_retail_int,
                "price_opt": price_opt_int,
            }
        )

    logger.info("Сток %s: собрано позиций: %d", code, len(rows))
    return json.dumps(rows, ensure_ascii=False, indent=2)


async def parse_feed_to_json(
    *,
    mode: Literal["catalog", "stock"] = "catalog",
    code: str = "D5",
    timeout: int = 30,
) -> str:
    """
    Унифицированная обёртка для D5.
    mode = 'catalog' -> Excel с Google Drive (папка из dropship_enterprises.gdrive_folder)
    mode = 'stock'   -> XML фид по dropship_enterprises.code
    """
    if mode == "catalog":
        return await parse_feed_catalog_to_json(code=code, timeout=timeout)
    elif mode == "stock":
        return await parse_feed_stock_to_json(code=code, timeout=timeout)
    else:
        raise ValueError("mode must be 'catalog' or 'stock'")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description=(
            "Парсер поставщика D5: режимы 'catalog' (Excel с Google Drive, "
            "папка из dropship_enterprises.gdrive_folder) "
            "и 'stock' (остатки/цены из XML фида). "
            "URL фида берётся из БД по dropship_enterprises.code"
        )
    )
    parser.add_argument(
        "--mode",
        choices=["catalog", "stock"],
        default="catalog",
        help="Режим обработки: catalog | stock (по умолчанию catalog)",
    )
    parser.add_argument(
        "--code",
        default="D5",
        help="значение поля code в dropship_enterprises (по умолчанию D5)",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=30,
        help="таймаут HTTP-запроса к фиду, сек. (для стока)",
    )

    args = parser.parse_args()
    out = asyncio.run(
        parse_feed_to_json(
            mode=args.mode,
            code=args.code,
            timeout=args.timeout,
        )
    )
    print(out)