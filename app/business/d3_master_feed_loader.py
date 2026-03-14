from __future__ import annotations

import argparse
import asyncio
import hashlib
import io
import json
import logging
import os
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

import httpx
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook
from sqlalchemy import select, text

from app.business.order_sender import SUPPLIERLIST_MAP
from app.database import get_async_db
from app.models import RawSupplierFeedProduct


logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger("d3_master_feed_loader")

D3_CODE = "D3"
EXCEL_COL_ARTIKUL = "Артикул"
EXCEL_COL_NAME = "Номенклатура"
EXCEL_COL_BARCODE = "Штрихкод"


@dataclass
class LoaderStats:
    supplier_id: int
    xml_items_read: int = 0
    excel_items_read: int = 0
    excel_matches_found: int = 0
    inserted: int = 0
    updated: int = 0
    warnings_count: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "supplier_id": self.supplier_id,
            "xml_items_read": self.xml_items_read,
            "excel_items_read": self.excel_items_read,
            "excel_matches_found": self.excel_matches_found,
            "inserted": self.inserted,
            "updated": self.updated,
            "warnings_count": self.warnings_count,
        }


def _warn(stats: LoaderStats, message: str, *args: Any) -> None:
    stats.warnings_count += 1
    logger.warning(message, *args)


def _normalize_string(value: Any) -> Optional[str]:
    if value is None:
        return None
    text_value = str(value).strip()
    return text_value or None


def _extract_supplier_id() -> int:
    supplier_token = SUPPLIERLIST_MAP.get(D3_CODE)
    if not supplier_token:
        raise RuntimeError(f"Не найден supplier mapping для {D3_CODE}")

    match = re.search(r"(\d+)$", supplier_token)
    if not match:
        raise RuntimeError(f"Не удалось извлечь supplier_id из значения {supplier_token!r} для {D3_CODE}")
    return int(match.group(1))


def _get_d3_folder_id() -> str:
    folder_id = (os.getenv("D3_CATALOG_FOLDER_ID") or "").strip()
    if not folder_id:
        raise RuntimeError("Не задана переменная окружения D3_CATALOG_FOLDER_ID")
    return folder_id


async def _connect_drive():
    creds_path = (os.getenv("GOOGLE_DRIVE_CREDENTIALS_PATH") or "").strip()
    if not creds_path or not os.path.exists(creds_path):
        raise FileNotFoundError(f"GOOGLE_DRIVE_CREDENTIALS_PATH не найден: {creds_path}")

    credentials = service_account.Credentials.from_service_account_file(
        creds_path,
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    return build("drive", "v3", credentials=credentials)


async def _fetch_single_file_metadata(drive_service, folder_id: str) -> Dict[str, Any]:
    try:
        results = (
            drive_service.files()
            .list(
                q=f"'{folder_id}' in parents and trashed=false",
                fields="files(id, name)",
                pageSize=10,
            )
            .execute()
        )
        files = results.get("files", []) or []
        if not files:
            raise FileNotFoundError(f"В папке {folder_id} нет файлов")
        return files[0]
    except HttpError as exc:
        raise RuntimeError(f"Ошибка Drive API при получении файла D3 каталога: {exc}") from exc


async def _download_file_bytes(drive_service, file_id: str) -> bytes:
    try:
        request = drive_service.files().get_media(fileId=file_id)
        handle = io.BytesIO()
        downloader = MediaIoBaseDownload(handle, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        handle.seek(0)
        return handle.read()
    except HttpError as exc:
        raise RuntimeError(f"Ошибка Drive API при загрузке файла {file_id}: {exc}") from exc


def _parse_excel_catalog(data: bytes) -> Dict[str, Dict[str, Optional[str]]]:
    workbook = load_workbook(io.BytesIO(data), data_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {
        str(value).strip(): idx
        for idx, value in enumerate(header_row)
        if value is not None
    }

    for column in (EXCEL_COL_ARTIKUL, EXCEL_COL_NAME, EXCEL_COL_BARCODE):
        if column not in headers:
            raise RuntimeError(f"В Excel D3 отсутствует обязательная колонка: {column}")

    idx_artikul = headers[EXCEL_COL_ARTIKUL]
    idx_name = headers[EXCEL_COL_NAME]
    idx_barcode = headers[EXCEL_COL_BARCODE]

    result: Dict[str, Dict[str, Optional[str]]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        artikul = _normalize_string(row[idx_artikul] if idx_artikul < len(row) else None)
        if not artikul:
            continue

        result[artikul] = {
            "excel_name": _normalize_string(row[idx_name] if idx_name < len(row) else None),
            "barcode": _normalize_string(row[idx_barcode] if idx_barcode < len(row) else None),
        }

    return result


async def _get_feed_url_by_code(code: str = D3_CODE) -> Optional[str]:
    async with get_async_db() as session:
        result = await session.execute(
            text("SELECT feed_url FROM dropship_enterprises WHERE code = :code LIMIT 1"),
            {"code": code},
        )
        return result.scalar_one_or_none()


async def _load_feed_root(code: str = D3_CODE, timeout: int = 30) -> ET.Element:
    feed_url = await _get_feed_url_by_code(code)
    if not feed_url:
        raise RuntimeError(f"Не найден feed_url в dropship_enterprises для code='{code}'")

    headers = {"User-Agent": "Mozilla/5.0"}
    async with httpx.AsyncClient(headers=headers, timeout=timeout) as client:
        response = await client.get(feed_url)
        response.raise_for_status()

    try:
        return ET.fromstring(response.text)
    except ET.ParseError as exc:
        raise RuntimeError(f"Ошибка парсинга XML фида D3: {exc}") from exc


def _collect_offers(root: ET.Element, limit: int = 0) -> List[ET.Element]:
    offers = root.findall(".//offer")
    if limit and limit > 0:
        return offers[:limit]
    return offers


def _get_text(node: ET.Element, tag: str) -> Optional[str]:
    child = node.find(tag)
    if child is None:
        return None
    return _normalize_string(child.text)


def _extract_images(offer: ET.Element) -> List[str]:
    result: List[str] = []
    seen = set()
    for child in offer.findall("picture"):
        image_url = _normalize_string(child.text)
        if not image_url or image_url in seen:
            continue
        seen.add(image_url)
        result.append(image_url)
    return result


def _extract_params(offer: ET.Element) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for param in offer.findall("param"):
        name = _normalize_string(param.get("name") or param.get("Name"))
        value = _normalize_string(param.text)
        if not name or not value:
            continue
        if name not in result:
            result[name] = value
    return result


def _build_source_hash(payload: Dict[str, Any]) -> str:
    serialized = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def _parse_offer(
    offer: ET.Element,
    excel_catalog: Dict[str, Dict[str, Optional[str]]],
    supplier_id: int,
    stats: LoaderStats,
) -> Optional[Dict[str, Any]]:
    supplier_code = _get_text(offer, "vendorCode") or _normalize_string(offer.get("vendorCode"))
    feed_product_id = _normalize_string(offer.get("id")) or _get_text(offer, "id")

    if not supplier_code:
        _warn(stats, "Пропущен D3 offer без vendorCode: offer_id=%r", feed_product_id)
        return None

    excel_item = excel_catalog.get(supplier_code)
    if excel_item is None:
        _warn(stats, "Не найдена запись в Excel D3 по vendorCode=%s", supplier_code)
    else:
        stats.excel_matches_found += 1

    barcode = excel_item.get("barcode") if excel_item else None
    if not barcode:
        _warn(stats, "Не найден barcode в Excel D3 для vendorCode=%s", supplier_code)

    name_ua = _get_text(offer, "name_ua")
    name_ru = _get_text(offer, "name")
    description_ua = _get_text(offer, "description_ua")
    description_ru = _get_text(offer, "description")
    images = _extract_images(offer)
    params = _extract_params(offer)
    excel_name = excel_item.get("excel_name") if excel_item else None

    name_raw = name_ua or name_ru or excel_name
    description_raw = description_ua or description_ru
    category_raw = _get_text(offer, "categoryId")

    source_payload = {
        "name_ua": name_ua,
        "name_ru": name_ru,
        "description_ua": description_ua,
        "description_ru": description_ru,
        "images": images,
        "params": params,
        "barcode_source": "excel_catalog" if barcode else None,
        "catalog_name_excel": excel_name,
        "barcode_excel": barcode,
        "category_id": category_raw,
        "offer_id": feed_product_id,
        "price": _get_text(offer, "price"),
        "priceRRP": _get_text(offer, "priceRRP"),
        "vendorPrice": _get_text(offer, "vendorPrice"),
        "currencyId": _get_text(offer, "currencyId"),
        "available": _normalize_string(offer.get("available")),
        "in_stock": _normalize_string(offer.get("in_stock")),
        "presence_sure": _normalize_string(offer.get("presence_sure")),
        "promotion": _normalize_string(offer.get("promotion")),
        "group_id": _normalize_string(offer.get("group_id")),
    }

    return {
        "supplier_id": supplier_id,
        "feed_product_id": feed_product_id,
        "supplier_code": supplier_code,
        "name_raw": name_raw,
        "manufacturer_raw": _get_text(offer, "vendor"),
        "barcode": barcode,
        "description_raw": description_raw,
        "category_raw": category_raw,
        "source_payload": source_payload,
        "source_hash": _build_source_hash(
            {
                "supplier_code": supplier_code,
                "barcode": barcode,
                "name_raw": name_raw,
                "manufacturer_raw": _get_text(offer, "vendor"),
                "description_raw": description_raw,
                "category_raw": category_raw,
                "images": images,
            }
        ),
    }


async def load_d3_raw_supplier_feed(limit: int = 0) -> Dict[str, Any]:
    supplier_id = _extract_supplier_id()
    stats = LoaderStats(supplier_id=supplier_id)
    logger.info("Запуск D3 master feed loader, supplier_id=%s", supplier_id)

    drive = await _connect_drive()
    folder_id = _get_d3_folder_id()
    file_meta = await _fetch_single_file_metadata(drive, folder_id)
    logger.info("Найден файл Excel D3: %s (%s)", file_meta.get("name"), file_meta.get("id"))
    file_bytes = await _download_file_bytes(drive, file_meta["id"])
    excel_catalog = _parse_excel_catalog(file_bytes)
    stats.excel_items_read = len(excel_catalog)

    root = await _load_feed_root(code=D3_CODE)
    offers = _collect_offers(root, limit=limit)
    stats.xml_items_read = len(offers)

    async with get_async_db() as session:
        for offer in offers:
            parsed = _parse_offer(offer, excel_catalog, supplier_id, stats)
            if parsed is None:
                continue

            stmt = select(RawSupplierFeedProduct).where(
                RawSupplierFeedProduct.supplier_id == supplier_id,
                RawSupplierFeedProduct.supplier_code == parsed["supplier_code"],
            )
            existing = (await session.execute(stmt)).scalar_one_or_none()

            if existing is None:
                session.add(
                    RawSupplierFeedProduct(
                        supplier_id=supplier_id,
                        feed_product_id=parsed["feed_product_id"],
                        supplier_code=parsed["supplier_code"],
                        name_raw=parsed["name_raw"],
                        manufacturer_raw=parsed["manufacturer_raw"],
                        barcode=parsed["barcode"],
                        description_raw=parsed["description_raw"],
                        weight_g=None,
                        length_mm=None,
                        width_mm=None,
                        height_mm=None,
                        volume_ml=None,
                        category_raw=parsed["category_raw"],
                        source_payload=parsed["source_payload"],
                        source_hash=parsed["source_hash"],
                    )
                )
                stats.inserted += 1
                continue

            existing.feed_product_id = parsed["feed_product_id"]
            existing.name_raw = parsed["name_raw"]
            existing.manufacturer_raw = parsed["manufacturer_raw"]
            existing.barcode = parsed["barcode"]
            existing.description_raw = parsed["description_raw"]
            existing.category_raw = parsed["category_raw"]
            existing.source_payload = parsed["source_payload"]
            existing.source_hash = parsed["source_hash"]
            stats.updated += 1

    logger.info(
        "D3 master feed loader завершён: xml_items_read=%d, excel_items_read=%d, inserted=%d, updated=%d",
        stats.xml_items_read,
        stats.excel_items_read,
        stats.inserted,
        stats.updated,
    )
    return stats.to_dict()


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Загрузка D3 в raw_supplier_feed_products для master-контура"
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="обработать только первые N XML offer (0 = без лимита)",
    )
    return parser.parse_args()


async def _amain() -> None:
    args = _parse_args()
    result = await load_d3_raw_supplier_feed(limit=args.limit)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    asyncio.run(_amain())
