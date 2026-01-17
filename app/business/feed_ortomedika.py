from __future__ import annotations

import asyncio
import io
import json
import logging
import os
from typing import Optional, List, Dict, Literal, Any, Tuple

import httpx
import xml.etree.ElementTree as ET
from sqlalchemy import text

from app.database import get_async_db
from app.services.notification_service import send_notification

# Google Drive
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload

# Excel
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


# === НАСТРОЙКИ D9 (ORTOMEDIKA) ===

D9_STOCK_FEEDS: List[Tuple[str, str]] = [
    (
        "Aurafix",
        "https://ortomedika.com.ua/products_feed.xml?hash_tag=63b804be40f92e4537e5d64245e3875b&sales_notes=&product_ids=&label_ids=10410446&exclude_fields=&html_description=1&yandex_cpa=&process_presence_sure=&languages=uk%2Cru&extra_fields=keywords&group_ids=",
    ),
    (
        "Thuasne",
        "https://ortomedika.com.ua/products_feed.xml?hash_tag=63b804be40f92e4537e5d64245e3875b&sales_notes=&product_ids=&label_ids=9563942%2C108276523&exclude_fields=&html_description=1&yandex_cpa=&process_presence_sure=&languages=uk%2Cru&extra_fields=keywords&group_ids=",
    ),
    (
        "Bergal",
        "https://ortomedika.com.ua/products_feed.xml?hash_tag=63b804be40f92e4537e5d64245e3875b&sales_notes=&product_ids=&label_ids=14270749&exclude_fields=&html_description=1&yandex_cpa=&process_presence_sure=&languages=uk%2Cru&extra_fields=keywords&group_ids=",
    ),
    (
        "Sporlastic",
        "https://ortomedika.com.ua/products_feed.xml?hash_tag=63b804be40f92e4537e5d64245e3875b&sales_notes=&product_ids=&label_ids=21495429&exclude_fields=&html_description=1&yandex_cpa=&process_presence_sure=&languages=uk%2Cru&extra_fields=keywords&group_ids=",
    ),
    (
        "Ortofix",
        "https://ortomedika.com.ua/products_feed.xml?hash_tag=63b804be40f92e4537e5d64245e3875b&sales_notes=&product_ids=&label_ids=9563739%2C70722672&exclude_fields=&html_description=1&yandex_cpa=&process_presence_sure=&languages=uk%2Cru&extra_fields=keywords&group_ids=",
    ),
    (
        "D3",
        "https://ortomedika.com.ua/products_feed.xml?hash_tag=63b804be40f92e4537e5d64245e3875b&sales_notes=&product_ids=&label_ids=112825345&exclude_fields=&html_description=1&yandex_cpa=&process_presence_sure=&languages=uk%2Cru&extra_fields=keywords&group_ids=",
    ),
    (
        "Ortomedica_no_shoes",
        "https://ortomedika.com.ua/products_feed.xml?hash_tag=63b804be40f92e4537e5d64245e3875b&sales_notes=&product_ids=&label_ids=107742428&exclude_fields=&html_description=1&yandex_cpa=&process_presence_sure=&languages=uk%2Cru&extra_fields=keywords&group_ids=",
    ),
]


# === ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ===

def _get_text(el: ET.Element, candidates: List[str]) -> Optional[str]:
    """Возвращает текст первого дочернего тега из списка кандидатов."""
    for tag in candidates:
        child = el.find(tag)
        if child is not None and child.text and child.text.strip():
            return child.text.strip()
    return None


def _to_float(val: Optional[str]) -> float:
    """Мягко преобразует строку в float (учитывая пробелы/запятые)."""
    if not val:
        return 0.0
    s = str(val).strip()
    s = s.replace(" ", "").replace("\u00A0", "")
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


async def _get_gdrive_folder_by_code(code: str) -> Optional[str]:
    """Достаёт gdrive_folder из dropship_enterprises по code."""
    async with get_async_db() as session:
        res = await session.execute(
            text("SELECT gdrive_folder FROM dropship_enterprises WHERE code = :code LIMIT 1"),
            {"code": code},
        )
        return res.scalar_one_or_none()


async def _get_profit_percent_by_code(code: str = "D1") -> float:
    """Достаёт profit_percent (%) из dropship_enterprises по значению поля code.

    В БД значение хранится как проценты (например, 25), возвращаем долю (0.25).
    """
    async with get_async_db() as session:
        res = await session.execute(
            text("SELECT profit_percent FROM dropship_enterprises WHERE code = :code LIMIT 1"),
            {"code": code},
        )
        raw = res.scalar_one_or_none()

    try:
        val = float(raw)
    except Exception:
        val = 0.0

    if val > 1:
        val = val / 100.0

    if val < 0:
        val = 0.0
    if val > 1:
        val = 1.0

    return val


def _collect_offer_nodes(root: ET.Element) -> List[ET.Element]:
    """Собираем узлы <offer> из фида."""
    offers = root.findall(".//offer")
    if not offers:
        offers = [el for el in root.iter() if el.tag.lower() in ("offer", "item")]
    return offers


# === GOOGLE DRIVE / EXCEL ДЛЯ КАТАЛОГА ===

async def _connect_to_google_drive():
    """
    Создает клиент Drive API через сервисный аккаунт.
    Использует GOOGLE_DRIVE_CREDENTIALS_PATH.
    """
    creds_path = os.getenv("GOOGLE_DRIVE_CREDENTIALS_PATH")
    if not creds_path or not os.path.exists(creds_path):
        msg = f"Неверный путь к учетным данным Google Drive: {creds_path}"
        logger.error(msg)
        send_notification(msg, "Разработчик")
        raise FileNotFoundError(msg)

    credentials = service_account.Credentials.from_service_account_file(
        creds_path,
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    service = build("drive", "v3", credentials=credentials)
    logger.info("Подключено к Google Drive")
    return service


async def _fetch_single_file_metadata(drive_service, folder_id: str) -> Dict[str, Any]:
    """
    Возвращает metadata ОДНОГО файла (id, name) из папки.
    По условию задачи в папке всегда один файл.
    """
    try:
        results = (
            drive_service.files()
            .list(
                q=f"'{folder_id}' in parents and trashed=false",
                fields="files(id, name)",
                pageSize=10,
            )
            .execute()
        )
        files = results.get("files", []) or []
        if not files:
            raise FileNotFoundError(f"В папке {folder_id} нет файлов")
        return files[0]
    except HttpError as e:
        msg = f"HTTP ошибка при получении файлов из папки {folder_id}: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        raise


async def _download_file_bytes(drive_service, file_id: str) -> bytes:
    """Загружает файл с Google Drive в bytes."""
    try:
        request = drive_service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)
        return fh.read()
    except HttpError as e:
        msg = f"HTTP ошибка при загрузке файла {file_id}: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        raise


def _parse_d9_catalog_excel_bytes(data: bytes) -> List[Dict[str, str]]:
    """
    Парсит Excel-файл из bytes и возвращает список:
    [{"id": ..., "name": ..., "barcode": ...}, ...]

    Маппинг D9:
      Код               -> id
      ТоварПредприятия  -> name
      ШК                -> barcode
    """
    wb = load_workbook(io.BytesIO(data), data_only=True)
    sheet = wb.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {str(value).strip(): idx for idx, value in enumerate(header_row) if value is not None}

    required_cols = ["Код", "ТоварПредприятия", "ШК"]
    for col in required_cols:
        if col not in headers:
            raise ValueError(f"В Excel не найден обязательный столбец '{col}'")

    idx_id = headers["Код"]
    idx_name = headers["ТоварПредприятия"]
    idx_barcode = headers["ШК"]

    items: List[Dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code_val = row[idx_id] if idx_id < len(row) else None
        name_val = row[idx_name] if idx_name < len(row) else None
        barcode_val = row[idx_barcode] if idx_barcode < len(row) else None

        if not code_val or not name_val:
            continue

        items.append(
            {
                "id": str(code_val).strip(),
                "name": str(name_val).strip(),
                "barcode": (str(barcode_val).strip() if barcode_val else ""),
            }
        )

    logger.info("Каталог (Excel D9): собрано позиций: %d", len(items))
    return items


# === XML ФИДЫ ДЛЯ СТОКА ===

async def _load_feed_root_from_url(*, url: str, timeout: int) -> Optional[ET.Element]:
    """Скачивает XML по URL и возвращает корень."""
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        async with httpx.AsyncClient(headers=headers, timeout=timeout) as client:
            resp = await client.get(url)
            resp.raise_for_status()
            xml_text = resp.text
    except Exception as e:
        msg = f"Ошибка загрузки фида {url}: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        return None

    try:
        return ET.fromstring(xml_text)
    except Exception as e:
        msg = f"Ошибка парсинга XML из {url}: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        return None


# === ОСНОВНЫЕ ФУНКЦИИ ===

async def parse_feed_catalog_to_json(
    *,
    code: str = "D9",
    timeout: int = 30,  # не используется, оставлено для совместимости
) -> str:
    """
    Каталог D9:
    1) gdrive_folder берём из dropship_enterprises по code
    2) в папке 1 Excel файл
    3) читаем столбцы:
        Код -> id
        ТоварПредприятия -> name
        ШК -> barcode
    """
    try:
        folder_id = await _get_gdrive_folder_by_code(code)
        if not folder_id:
            raise RuntimeError(f"Не найден gdrive_folder в dropship_enterprises для code='{code}'")

        drive_service = await _connect_to_google_drive()
        file_meta = await _fetch_single_file_metadata(drive_service, folder_id)
        file_id = file_meta["id"]
        file_name = file_meta.get("name")
        logger.info("Найден файл каталога для %s: %s (%s)", code, file_name, file_id)

        file_bytes = await _download_file_bytes(drive_service, file_id)
        items = _parse_d9_catalog_excel_bytes(file_bytes)
    except Exception as e:
        msg = f"Ошибка обработки каталога {code} из Google Drive: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        return "[]"

    return json.dumps(items, ensure_ascii=False, indent=2)


async def parse_feed_stock_to_json(
    *,
    code: str = "D9",
    timeout: int = 30,
) -> str:
    """
    Сток D9: обходит несколько XML фидов (захардкожены) и склеивает в один JSON:
    [
      {"code_sup": "<vendorCode>", "qty": 1, "price_retail": <int>, "price_opt": <float>},
      ...
    ]

    Правила:
      - берём только offer с available="true"
      - vendorCode -> code_sup
      - price -> price_retail
      - price_opt = price_retail / (1 + profit_percent), profit_percent из dropship_enterprises по code
    """
    profit_percent = await _get_profit_percent_by_code(code)

    # На случай дублей vendorCode между фидами:
    # сохраняем последнюю встреченную запись (можно легко поменять стратегию при необходимости)
    merged: Dict[str, Dict[str, Any]] = {}

    for feed_name, feed_url in D9_STOCK_FEEDS:
        root = await _load_feed_root_from_url(url=feed_url, timeout=timeout)
        if root is None:
            logger.warning("Пропускаю фид %s из-за ошибки загрузки/парсинга", feed_name)
            continue

        count_added = 0
        for offer in _collect_offer_nodes(root):
            available_raw = (offer.get("available") or "").strip().lower()
            if available_raw != "true":
                continue

            vendor_code = _get_text(offer, ["vendorCode"]) or offer.get("vendorCode")
            if not vendor_code:
                continue

            price_raw = _get_text(offer, ["price"])
            price_retail = _to_float(price_raw)
            price_retail_int = int(price_retail)

            qty = 1  # строго по ТЗ

            price_opt = price_retail / (1.0 + profit_percent) if (1.0 + profit_percent) != 0 else 0.0
            if price_opt < 0:
                price_opt = 0.0
            price_opt = round(price_opt, 2)

            merged[str(vendor_code).strip()] = {
                "code_sup": str(vendor_code).strip(),
                "qty": qty,
                "price_retail": price_retail_int,
                "price_opt": price_opt,
            }
            count_added += 1

        logger.info("Сток D9: фид '%s' обработан, добавлено/обновлено: %d", feed_name, count_added)

    rows = list(merged.values())
    logger.info("Сток D9: итоговых уникальных позиций (code=%s): %d", code, len(rows))
    return json.dumps(rows, ensure_ascii=False, indent=2)


async def parse_feed_to_json(
    *,
    mode: Literal["catalog", "stock"] = "catalog",
    code: str = "D9",
    timeout: int = 30,
) -> str:
    """Унифицированная обёртка для D9."""
    if mode == "catalog":
        return await parse_feed_catalog_to_json(code=code, timeout=timeout)
    if mode == "stock":
        return await parse_feed_stock_to_json(code=code, timeout=timeout)
    raise ValueError("mode must be 'catalog' or 'stock'")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description=(
            "Парсер поставщика D9 (Ortomedika): "
            "режим 'catalog' (Excel из Google Drive по dropship_enterprises.gdrive_folder) "
            "и режим 'stock' (остатки/цены из нескольких XML фидов по захардкоженным URL)."
        )
    )
    parser.add_argument(
        "--mode",
        choices=["catalog", "stock"],
        default="catalog",
        help="Режим обработки: catalog | stock (по умолчанию catalog)",
    )
    parser.add_argument(
        "--code",
        default="D9",
        help="значение поля code в dropship_enterprises (по умолчанию D9)",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=30,
        help="таймаут HTTP-запросов, сек. (по умолчанию 30)",
    )

    args = parser.parse_args()
    out = asyncio.run(parse_feed_to_json(mode=args.mode, code=args.code, timeout=args.timeout))
    print(out)