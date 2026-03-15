from __future__ import annotations

import asyncio
import io
import json
import logging
import os
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional, Tuple

import httpx
import xml.etree.ElementTree as ET
from sqlalchemy import text

from app.database import get_async_db
from app.services.notification_service import send_notification

# Excel
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# === D10 (ZooHub) ===
D10_CODE_DEFAULT = "D10"
D10_PRICE_URL_DEFAULT = "https://zoohub.ua/c_integr/dilovod/priceList.xlsx"

# Excel: заголовки на 4-й строке (1-indexed)
D10_HEADER_ROW = 4
D10_HEADER_SEARCH_MAX_ROWS = 20

# Кэш/стейт (для дельты и drop-цен)
# Можно переопределить env: D10_STATE_DIR=/path/to/state
DEFAULT_STATE_DIR = Path(os.getenv("D10_STATE_DIR") or "state_cache")


# ──────────────────────────────────────────────────────────────────────────────
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ──────────────────────────────────────────────────────────────────────────────

def _strip_ns(tag: str) -> str:
    if not tag:
        return tag
    return tag.split("}")[-1]


def _to_float(val: Optional[str]) -> float:
    if val is None:
        return 0.0
    s = str(val).strip()
    if not s:
        return 0.0
    # Уберём валюты/текст, оставим цифры/точку/запятую/минус
    s = s.replace("\u00A0", " ").strip()
    # Частый кейс: "235 UAH"
    m = re.search(r"-?\d+(?:[.,]\d+)?", s)
    if not m:
        return 0.0
    num = m.group(0).replace(",", ".")
    try:
        return float(num)
    except Exception:
        return 0.0


def _norm_str(v: Any) -> str:
    s = ("" if v is None else str(v)).strip()
    # иногда Excel даёт "None" текстом — прибьём
    if s.lower() == "none":
        return ""
    return s


def _is_probably_junk_row(values: Dict[str, str]) -> bool:
    """
    Пропускаем строки-категории/мусор:
    - нет артикула
    - или заполнено только одно поле (например, "антибиотики"), без ШК и цены
    """
    art = values.get("id", "")
    name = values.get("name", "")
    barcode = values.get("barcode", "")
    drop_price = values.get("drop_price_raw", "")

    if not art:
        return True

    # если “артикул” похож на текстовую категорию, иногда так бывает — отсеем
    # (если артикул не содержит цифр и слишком длинный)
    if not re.search(r"\d", art) and len(art) > 12:
        return True

    # если это “раздел”, часто: есть только name, а barcode/цены пустые
    if name and not barcode and not drop_price and len(name) > 2:
        # но если есть артикул — всё же может быть товар; оставим мягко:
        # считаем мусором только если name == art (часто у категорий)
        if name.strip().lower() == art.strip().lower():
            return True

    return False


# ──────────────────────────────────────────────────────────────────────────────
# DB HELPERS
# ──────────────────────────────────────────────────────────────────────────────

async def _get_feed_url_by_code(code: str) -> Optional[str]:
    async with get_async_db() as session:
        res = await session.execute(
            text("SELECT feed_url FROM dropship_enterprises WHERE code = :code LIMIT 1"),
            {"code": code},
        )
        return res.scalar_one_or_none()


async def _get_profit_percent_by_code(code: str) -> Optional[float]:
    async with get_async_db() as session:
        res = await session.execute(
            text("SELECT profit_percent FROM dropship_enterprises WHERE code = :code LIMIT 1"),
            {"code": code},
        )
        value = res.scalar_one_or_none()
        if value is None:
            return None
        try:
            return float(value)
        except Exception:
            return None


def _get_price_url() -> str:
    return (os.getenv("ZOOHUB_PRICE_URL") or D10_PRICE_URL_DEFAULT).strip()


async def _download_excel_bytes_from_url(url: str, timeout: int = 60) -> bytes:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*;q=0.8",
    }
    try:
        async with httpx.AsyncClient(headers=headers, timeout=timeout, follow_redirects=True) as client:
            resp = await client.get(url)
            resp.raise_for_status()
            return resp.content
    except Exception as e:
        msg = f"D10: Ошибка загрузки Excel по URL {url}: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        raise


async def _load_catalog_items_from_excel_url(*, timeout: int) -> List[Dict[str, Any]]:
    price_url = _get_price_url()
    logger.info("D10: Загрузка Excel из URL: %s", price_url)
    file_bytes = await _download_excel_bytes_from_url(price_url, timeout=timeout)
    return _parse_d10_catalog_excel_bytes(file_bytes)


# ──────────────────────────────────────────────────────────────────────────────
# STATE CACHE (для дельты и drop-цен)
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class D10State:
    # ключ: id (артикул) -> signature
    sig_by_id: Dict[str, str]
    # drop price cache
    drop_by_id: Dict[str, float]
    drop_by_barcode: Dict[str, float]

    @classmethod
    def empty(cls) -> "D10State":
        return cls(sig_by_id={}, drop_by_id={}, drop_by_barcode={})

    @staticmethod
    def _state_paths(code: str) -> Tuple[Path, Path]:
        DEFAULT_STATE_DIR.mkdir(parents=True, exist_ok=True)
        state_path = DEFAULT_STATE_DIR / f"{code.lower()}_catalog_state.json"
        drop_path = DEFAULT_STATE_DIR / f"{code.lower()}_drop_prices.json"
        return state_path, drop_path

    @classmethod
    def load(cls, code: str) -> "D10State":
        state_path, drop_path = cls._state_paths(code)

        sig_by_id: Dict[str, str] = {}
        drop_by_id: Dict[str, float] = {}
        drop_by_barcode: Dict[str, float] = {}

        if state_path.exists():
            try:
                sig_by_id = json.loads(state_path.read_text(encoding="utf-8"))
                if not isinstance(sig_by_id, dict):
                    sig_by_id = {}
            except Exception:
                sig_by_id = {}

        if drop_path.exists():
            try:
                payload = json.loads(drop_path.read_text(encoding="utf-8"))
                if isinstance(payload, dict):
                    drop_by_id = payload.get("drop_by_id") or {}
                    drop_by_barcode = payload.get("drop_by_barcode") or {}
                    # нормализация типов
                    drop_by_id = {str(k): float(v) for k, v in drop_by_id.items() if str(k).strip()}
                    drop_by_barcode = {str(k): float(v) for k, v in drop_by_barcode.items() if str(k).strip()}
            except Exception:
                drop_by_id = {}
                drop_by_barcode = {}

        return cls(sig_by_id=sig_by_id, drop_by_id=drop_by_id, drop_by_barcode=drop_by_barcode)

    def save(self, code: str) -> None:
        state_path, drop_path = self._state_paths(code)

        tmp_state = state_path.with_suffix(".json.partial")
        tmp_drop = drop_path.with_suffix(".json.partial")

        tmp_state.write_text(json.dumps(self.sig_by_id, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp_drop.write_text(
            json.dumps(
                {"drop_by_id": self.drop_by_id, "drop_by_barcode": self.drop_by_barcode},
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        os.replace(tmp_state, state_path)
        os.replace(tmp_drop, drop_path)


def _make_signature(*, item_id: str, name: str, barcode: str, drop_price: float) -> str:
    # сигнатура должна меняться при любом значимом изменении
    return f"{item_id}|{name}|{barcode}|{drop_price:.4f}"


# ──────────────────────────────────────────────────────────────────────────────
# CATALOG: Excel parsing (строка заголовков = 4)
# ──────────────────────────────────────────────────────────────────────────────

D10_HEADER_ALIASES = {
    "id": [
        "Артикул",
        "Артикул товару",
        "Артикул товара",
        "Код",
        "Код товару",
        "Код товара",
    ],
    "name": [
        "Найменування товару",
        "Найменування",
        "Назва товару",
        "Назва",
    ],
    "barcode": [
        "Штрихкод",
        "Штрих-код",
        "EAN",
        "EAN-13",
    ],
    "drop_price_raw": [
        "Ціна – drop",
        "Ціна - drop",
        "Ціна-drop",
        "drop, грн.",
        "Drop",
    ],
}


def _normalize_header_text(value: Any) -> str:
    s = _norm_str(value).lower()
    if not s:
        return ""
    s = s.replace("\u00A0", " ")
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _resolve_header_mapping(sheet: Any) -> Tuple[int, Dict[str, int]]:
    """
    Ищет строку заголовков в верхней части листа и возвращает:
    (header_row_idx, {"id": col_idx, "name": col_idx, "barcode": col_idx, "drop_price_raw": col_idx})
    """
    alias_norm: Dict[str, List[str]] = {
        field: [_normalize_header_text(a) for a in aliases if _normalize_header_text(a)]
        for field, aliases in D10_HEADER_ALIASES.items()
    }

    best_row = D10_HEADER_ROW
    best_score = -1
    best_missing: List[str] = list(D10_HEADER_ALIASES.keys())

    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=D10_HEADER_SEARCH_MAX_ROWS, values_only=True),
        start=1,
    ):
        headers_norm: Dict[str, int] = {}
        for idx, value in enumerate(row):
            key = _normalize_header_text(value)
            if key:
                headers_norm[key] = idx

        idx_map: Dict[str, int] = {}
        missing: List[str] = []
        for field_name, aliases in alias_norm.items():
            found_idx: Optional[int] = None
            for alias in aliases:
                if alias in headers_norm:
                    found_idx = headers_norm[alias]
                    break
            if found_idx is None:
                missing.append(field_name)
            else:
                idx_map[field_name] = found_idx

        score = len(D10_HEADER_ALIASES) - len(missing)
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_missing = missing

        if not missing:
            logger.info("D10: Header row auto-detected at row=%d", row_idx)
            return row_idx, idx_map

    raise ValueError(
        "D10: Не удалось определить заголовки Excel. "
        f"Лучшее совпадение: row={best_row}, найдено={best_score}/{len(D10_HEADER_ALIASES)}, "
        f"не хватает={best_missing}"
    )


def _normalize_barcode(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)

    raw = str(value).strip().replace("\u00A0", "").replace(" ", "")
    if not raw:
        return ""

    # Numeric formats (including scientific notation) normalize via Decimal.
    if re.fullmatch(r"[+-]?\d+(?:[.,]\d+)?(?:[eE][+-]?\d+)?", raw):
        normalized_num = raw.replace(",", ".")
        try:
            dec = Decimal(normalized_num)
        except InvalidOperation:
            return raw
        if dec.is_nan() or dec.is_infinite():
            return raw
        as_fixed = format(dec, "f")
        if "." in as_fixed:
            as_fixed = as_fixed.rstrip("0").rstrip(".")
        return as_fixed

    return raw

def _parse_d10_catalog_excel_bytes(data: bytes) -> List[Dict[str, Any]]:
    """
    Возвращает список записей:
    [
      {"id": "...", "name": "...", "barcode": "...", "drop_price": 12.34},
      ...
    ]
    """
    wb = load_workbook(io.BytesIO(data), data_only=True)
    sheet = wb.active

    header_row_idx, idx_map = _resolve_header_mapping(sheet)

    items: List[Dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        raw_id = row[idx_map["id"]] if idx_map["id"] < len(row) else None
        raw_name = row[idx_map["name"]] if idx_map["name"] < len(row) else None
        raw_barcode = row[idx_map["barcode"]] if idx_map["barcode"] < len(row) else None
        raw_drop = row[idx_map["drop_price_raw"]] if idx_map["drop_price_raw"] < len(row) else None

        item_id = _norm_str(raw_id)
        name = _norm_str(raw_name)
        barcode = _normalize_barcode(raw_barcode)
        drop_price = _to_float(_norm_str(raw_drop))

        row_obj = {
            "id": item_id,
            "name": name,
            "barcode": barcode,
            "drop_price": drop_price,
            "drop_price_raw": _norm_str(raw_drop),
        }

        if _is_probably_junk_row(row_obj):
            continue

        # Минимальная валидация: должен быть артикул и название (по опыту)
        if not item_id or not name:
            continue

        items.append(row_obj)

    logger.info("D10: Каталог Excel распарсен, позиций: %d", len(items))
    return items


def _apply_delta_and_update_state(
    *,
    code: str,
    items: List[Dict[str, Any]],
) -> List[Dict[str, str]]:
    """
    Возвращает только изменившиеся позиции для передачи в import_catalog.py:
      [{"id": ..., "name": ..., "barcode": ...}, ...]

    И обновляет стейт:
      - sig_by_id
      - drop_by_id, drop_by_barcode
    """
    state = D10State.load(code)
    is_first_run = len(state.sig_by_id) == 0

    changed: List[Dict[str, str]] = []

    new_sig_by_id: Dict[str, str] = dict(state.sig_by_id)
    new_drop_by_id: Dict[str, float] = dict(state.drop_by_id)
    new_drop_by_barcode: Dict[str, float] = dict(state.drop_by_barcode)

    for it in items:
        item_id = _norm_str(it.get("id"))
        name = _norm_str(it.get("name"))
        barcode = _norm_str(it.get("barcode"))
        drop_price = float(it.get("drop_price") or 0.0)

        sig = _make_signature(item_id=item_id, name=name, barcode=barcode, drop_price=drop_price)
        prev_sig = state.sig_by_id.get(item_id)

        # первый запуск: отдаём всё
        if is_first_run or (prev_sig != sig):
            changed.append({"id": item_id, "name": name, "barcode": barcode})

        # обновляем стейт
        new_sig_by_id[item_id] = sig
        new_drop_by_id[item_id] = drop_price
        if barcode:
            new_drop_by_barcode[barcode] = drop_price

    # сохраняем
    D10State(sig_by_id=new_sig_by_id, drop_by_id=new_drop_by_id, drop_by_barcode=new_drop_by_barcode).save(code)

    logger.info(
        "D10: Delta: first_run=%s, total=%d, changed=%d",
        is_first_run,
        len(items),
        len(changed),
    )
    return changed


async def _refresh_drop_cache_from_excel(*, code: str, timeout: int) -> D10State:
    """
    Обновляет drop-кэш из Excel URL и сохраняет в state_cache.
    Сигнатуры каталога (sig_by_id) не трогаем, чтобы не ломать delta-catalog поток.
    """
    state = D10State.load(code)
    new_drop_by_id, new_drop_by_barcode = await _load_drop_price_maps_from_excel_url(timeout=timeout)

    refreshed = D10State(
        sig_by_id=state.sig_by_id,
        drop_by_id=new_drop_by_id,
        drop_by_barcode=new_drop_by_barcode,
    )
    refreshed.save(code)
    logger.info(
        "D10: Drop cache refreshed from Excel URL, by_id=%d by_barcode=%d",
        len(refreshed.drop_by_id),
        len(refreshed.drop_by_barcode),
    )
    return refreshed


async def _load_drop_price_maps_from_excel_url(*, timeout: int) -> Tuple[Dict[str, float], Dict[str, float]]:
    """
    Автономно загружает Excel priceList.xlsx и строит drop price maps для stock flow.
    Не требует catalog state и не зависит от предыдущего catalog run.
    """
    logger.info("D10 stock: loading drop prices from Excel URL")
    items = await _load_catalog_items_from_excel_url(timeout=timeout)

    drop_by_id: Dict[str, float] = {}
    drop_by_barcode: Dict[str, float] = {}
    for it in items:
        item_id = _norm_str(it.get("id"))
        barcode = _norm_str(it.get("barcode"))
        drop_price = float(it.get("drop_price") or 0.0)
        if not item_id or drop_price <= 0.0:
            continue
        drop_by_id[item_id] = drop_price
        if barcode:
            drop_by_barcode[barcode] = drop_price

    logger.info(
        "D10 stock: drop prices loaded from Excel, by_id=%d by_barcode=%d",
        len(drop_by_id),
        len(drop_by_barcode),
    )
    return drop_by_id, drop_by_barcode


# ──────────────────────────────────────────────────────────────────────────────
# STOCK: XML feed from dropship_enterprises.feed_url
# ──────────────────────────────────────────────────────────────────────────────

async def _download_xml_feed(*, url: str, timeout: int) -> Optional[bytes]:
    headers = {"User-Agent": "Mozilla/5.0", "accept": "application/xml,text/xml,*/*"}
    try:
        async with httpx.AsyncClient(headers=headers, timeout=timeout) as client:
            resp = await client.get(url)
            resp.raise_for_status()
            return resp.content
    except Exception as e:
        msg = f"D10: Ошибка загрузки XML фида {url}: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        return None


def _find_child_text(elem: ET.Element, wanted_local_names: List[str]) -> str:
    """
    Ищет текст дочернего элемента по localname (без namespace), например:
      wanted_local_names=["mpn"] найдёт <g:mpn>...</g:mpn>
    """
    for ch in list(elem):
        if _strip_ns(ch.tag) in wanted_local_names:
            return (ch.text or "").strip()
    return ""


def _collect_item_nodes(root: ET.Element) -> List[ET.Element]:
    items = root.findall(".//item")
    if items:
        return items
    # fallback: любые item/offer
    return [el for el in root.iter() if _strip_ns(el.tag).lower() in ("item", "offer")]


def _get_drop_price_for_code_sup(
    drop_by_id: Dict[str, float],
    drop_by_barcode: Dict[str, float],
    code_sup: str,
) -> float:
    code_sup = (code_sup or "").strip()
    if not code_sup:
        return 0.0
    # основной кейс: g:mpn совпадает с Артикул
    if code_sup in drop_by_id:
        return float(drop_by_id.get(code_sup) or 0.0)
    # fallback: иногда mpn может совпасть с barcode (редко, но пусть будет)
    if code_sup in drop_by_barcode:
        return float(drop_by_barcode.get(code_sup) or 0.0)
    return 0.0


async def parse_feed_stock_to_json(
    *,
    code: str = D10_CODE_DEFAULT,
    timeout: int = 60,
) -> str:
    """
    Сток D10:
      - feed_url берём из dropship_enterprises.feed_url
      - парсим <item>
      - g:mpn -> code_sup
      - g:availability == "in stock" -> qty=1 (иначе пропускаем)
      - g:price -> price_retail (int)
      - price_opt -> из Excel drop price, грн. (по code_sup как по Артикулу)
    """
    feed_url = await _get_feed_url_by_code(code)
    if not feed_url:
        msg = f"D10: Не найден feed_url в dropship_enterprises для code='{code}'"
        logger.error(msg)
        send_notification(msg, "Разработчик")
        return "[]"

    xml_bytes = await _download_xml_feed(url=feed_url, timeout=timeout)
    if not xml_bytes:
        return "[]"

    try:
        root = ET.fromstring(xml_bytes)
    except Exception as e:
        msg = f"D10: Ошибка парсинга XML: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        return "[]"

    try:
        drop_by_id, drop_by_barcode = await _load_drop_price_maps_from_excel_url(timeout=timeout)
        try:
            refreshed_state = D10State.load(code)
            D10State(
                sig_by_id=refreshed_state.sig_by_id,
                drop_by_id=drop_by_id,
                drop_by_barcode=drop_by_barcode,
            ).save(code)
        except Exception as e:
            logger.warning("D10: Failed to persist stock drop cache to state_cache. error=%s", e)
    except Exception as e:
        logger.warning(
            "D10: Failed to load drop prices from Excel in stock flow, using fallback via existing cache if available. error=%s",
            e,
        )
        cached_state = D10State.load(code)
        drop_by_id = cached_state.drop_by_id
        drop_by_barcode = cached_state.drop_by_barcode
    profit_percent = await _get_profit_percent_by_code(code)
    profit_percent_dec = (float(profit_percent) / 100.0) if profit_percent is not None else 0.0

    rows: List[Dict[str, Any]] = []
    items = _collect_item_nodes(root)
    fallback_count = 0

    for it in items:
        # availability
        availability = _find_child_text(it, ["availability"]).strip().lower()
        if availability != "in stock":
            continue

        code_sup = _find_child_text(it, ["mpn"]).strip()
        if not code_sup:
            # fallback: иногда могут класть в g:id, но по ТЗ нужен mpn
            code_sup = _find_child_text(it, ["id"]).strip()
        if not code_sup:
            continue

        price_raw = _find_child_text(it, ["price"])
        price_retail = int(_to_float(price_raw))

        qty = 1  # строго по ТЗ
        price_opt = _get_drop_price_for_code_sup(drop_by_id, drop_by_barcode, code_sup)
        if price_opt <= 0.0 and profit_percent_dec > 0.0 and price_retail > 0:
            fallback_count += 1
            price_opt = float(price_retail) / (1.0 + profit_percent_dec)
        price_opt = round(float(price_opt), 2)

        rows.append(
            {
                "code_sup": code_sup,
                "qty": qty,
                "price_retail": price_retail,
                "price_opt": price_opt,
            }
        )

    logger.info("D10: Сток распарсен, позиций in stock: %d", len(rows))
    if fallback_count > 0:
        logger.warning(
            "D10: fallback via profit_percent applied for %d positions (qty > 0)",
            fallback_count,
        )
    return json.dumps(rows, ensure_ascii=False, indent=2)


# ──────────────────────────────────────────────────────────────────────────────
# CATALOG: Excel URL -> delta JSON
# ──────────────────────────────────────────────────────────────────────────────

async def parse_feed_catalog_to_json(
    *,
    code: str = D10_CODE_DEFAULT,
    timeout: int = 30,  # совместимость
) -> str:
    """
    Каталог D10:
      1) Скачиваем Excel по URL (env ZOOHUB_PRICE_URL)
      2) Заголовки в строке 4
      3) Выдаём: первый раз все, далее только изменения
      4) Сохраняем drop-price в локальный кэш для стока
    """
    try:
        all_items = await _load_catalog_items_from_excel_url(timeout=timeout)

        changed_items = _apply_delta_and_update_state(code=code, items=all_items)

        return json.dumps(changed_items, ensure_ascii=False, indent=2)

    except Exception as e:
        msg = f"D10: Ошибка обработки каталога из URL: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        return "[]"


async def parse_feed_to_json(
    *,
    mode: Literal["catalog", "stock"] = "catalog",
    code: str = D10_CODE_DEFAULT,
    timeout: int = 60,
) -> str:
    if mode == "catalog":
        return await parse_feed_catalog_to_json(code=code, timeout=timeout)
    if mode == "stock":
        return await parse_feed_stock_to_json(code=code, timeout=timeout)
    raise ValueError("mode must be 'catalog' or 'stock'")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description=(
            "Парсер поставщика D10 (ZooHub): "
            "catalog (Excel по URL, заголовки в строке 4, first full then delta, cache drop price) "
            "и stock (XML feed_url из dropship_enterprises, mpn->code_sup, price_opt from cached drop)."
        )
    )
    parser.add_argument("--mode", choices=["catalog", "stock"], default="catalog")
    parser.add_argument("--code", default=D10_CODE_DEFAULT)
    parser.add_argument("--timeout", type=int, default=60)

    args = parser.parse_args()
    out = asyncio.run(parse_feed_to_json(mode=args.mode, code=args.code, timeout=args.timeout))
    print(out)
