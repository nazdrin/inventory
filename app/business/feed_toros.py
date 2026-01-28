from __future__ import annotations

import asyncio
import io
import json
import logging
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional, Tuple

from sqlalchemy import text

from app.database import get_async_db
from app.services.notification_service import send_notification

# Google Drive
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload

# Excel xlsx
from openpyxl import load_workbook


logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# === D11 (Toros) ===
D11_CODE_DEFAULT = "D11"

# Кэш/стейт (для дельты + цен)
# Можно переопределить env: D11_STATE_DIR=/path/to/state
DEFAULT_STATE_DIR = Path(os.getenv("D11_STATE_DIR") or "state_cache")


# ──────────────────────────────────────────────────────────────────────────────
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ──────────────────────────────────────────────────────────────────────────────

def _to_float(val: Any) -> float:
    """
    Забирает число из строк типа "123,45 грн." / "123.45 UAH" / 123.45
    """
    if val is None:
        return 0.0
    s = str(val).replace("\u00A0", " ").strip()
    if not s:
        return 0.0
    m = re.search(r"-?\d+(?:[.,]\d+)?", s)
    if not m:
        return 0.0
    num = m.group(0).replace(",", ".")
    try:
        return float(num)
    except Exception:
        return 0.0


def _norm_str(v: Any) -> str:
    s = ("" if v is None else str(v)).strip()
    if s.lower() == "none":
        return ""
    return s


def _is_header_like(value: str) -> bool:
    v = (value or "").strip().lower()
    return v in {"артикул", "найменування", "штрихкод", "кількість", "количество"}


def _safe_int(v: Any) -> int:
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except Exception:
            return 0
    s = str(v).strip()
    if not s:
        return 0
    # иногда бывает "12,0"
    try:
        return int(float(s.replace(",", ".")))
    except Exception:
        # иногда "12 шт"
        m = re.search(r"-?\d+", s)
        return int(m.group(0)) if m else 0


# ──────────────────────────────────────────────────────────────────────────────
# DB HELPERS
# ──────────────────────────────────────────────────────────────────────────────

async def _get_gdrive_folder_by_code(code: str) -> Optional[str]:
    async with get_async_db() as session:
        res = await session.execute(
            text("SELECT gdrive_folder FROM dropship_enterprises WHERE code = :code LIMIT 1"),
            {"code": code},
        )
        return res.scalar_one_or_none()


async def _get_feed_url_by_code(code: str) -> Optional[str]:
    """
    По задаче: для D11 здесь хранится ID папки Google Drive для остатков.
    """
    async with get_async_db() as session:
        res = await session.execute(
            text("SELECT feed_url FROM dropship_enterprises WHERE code = :code LIMIT 1"),
            {"code": code},
        )
        return res.scalar_one_or_none()


# ──────────────────────────────────────────────────────────────────────────────
# GOOGLE DRIVE
# ──────────────────────────────────────────────────────────────────────────────

async def _connect_to_google_drive():
    creds_path = os.getenv("GOOGLE_DRIVE_CREDENTIALS_PATH")
    if not creds_path or not os.path.exists(creds_path):
        msg = f"D11: Неверный путь к учетным данным Google Drive: {creds_path}"
        logger.error(msg)
        send_notification(msg, "Разработчик")
        raise FileNotFoundError(msg)

    credentials = service_account.Credentials.from_service_account_file(
        creds_path,
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    service = build("drive", "v3", credentials=credentials)
    logger.info("D11: Подключено к Google Drive")
    return service


async def _fetch_latest_file_metadata(drive_service, folder_id: str) -> Dict[str, Any]:
    """
    Берём самый свежий файл в папке (по modifiedTime).
    """
    try:
        results = (
            drive_service.files()
            .list(
                q=f"'{folder_id}' in parents and trashed=false",
                fields="files(id, name, modifiedTime, mimeType)",
                orderBy="modifiedTime desc",
                pageSize=1,
            )
            .execute()
        )
        files = results.get("files", []) or []
        if not files:
            raise FileNotFoundError(f"D11: В папке {folder_id} нет файлов")
        return files[0]
    except HttpError as e:
        msg = f"D11: HTTP ошибка при получении файлов из папки {folder_id}: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        raise


async def _download_file_bytes(drive_service, file_id: str) -> bytes:
    try:
        request = drive_service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)
        return fh.read()
    except HttpError as e:
        msg = f"D11: HTTP ошибка при загрузке файла {file_id}: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        raise


# ──────────────────────────────────────────────────────────────────────────────
# STATE CACHE (дельта + цены)
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class D11State:
    # ключ: id (артикул) -> signature
    sig_by_id: Dict[str, str]
    # цены
    opt_by_id: Dict[str, float]
    retail_by_id: Dict[str, float]

    @classmethod
    def empty(cls) -> "D11State":
        return cls(sig_by_id={}, opt_by_id={}, retail_by_id={})

    @staticmethod
    def _state_paths(code: str) -> Tuple[Path, Path]:
        DEFAULT_STATE_DIR.mkdir(parents=True, exist_ok=True)
        state_path = DEFAULT_STATE_DIR / f"{code.lower()}_catalog_state.json"
        price_path = DEFAULT_STATE_DIR / f"{code.lower()}_prices.json"
        return state_path, price_path

    @classmethod
    def load(cls, code: str) -> "D11State":
        state_path, price_path = cls._state_paths(code)

        sig_by_id: Dict[str, str] = {}
        opt_by_id: Dict[str, float] = {}
        retail_by_id: Dict[str, float] = {}

        if state_path.exists():
            try:
                sig_by_id = json.loads(state_path.read_text(encoding="utf-8"))
                if not isinstance(sig_by_id, dict):
                    sig_by_id = {}
            except Exception:
                sig_by_id = {}

        if price_path.exists():
            try:
                payload = json.loads(price_path.read_text(encoding="utf-8"))
                if isinstance(payload, dict):
                    opt_by_id = payload.get("opt_by_id") or {}
                    retail_by_id = payload.get("retail_by_id") or {}
                    opt_by_id = {str(k): float(v) for k, v in opt_by_id.items() if str(k).strip()}
                    retail_by_id = {str(k): float(v) for k, v in retail_by_id.items() if str(k).strip()}
            except Exception:
                opt_by_id = {}
                retail_by_id = {}

        return cls(sig_by_id=sig_by_id, opt_by_id=opt_by_id, retail_by_id=retail_by_id)

    def save(self, code: str) -> None:
        state_path, price_path = self._state_paths(code)

        tmp_state = state_path.with_suffix(".json.partial")
        tmp_price = price_path.with_suffix(".json.partial")

        tmp_state.write_text(json.dumps(self.sig_by_id, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp_price.write_text(
            json.dumps(
                {"opt_by_id": self.opt_by_id, "retail_by_id": self.retail_by_id},
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        os.replace(tmp_state, state_path)
        os.replace(tmp_price, price_path)


def _make_signature(*, item_id: str, name: str, barcode: str, price_opt: float, price_retail: float) -> str:
    return f"{item_id}|{name}|{barcode}|{price_opt:.4f}|{price_retail:.4f}"


# ──────────────────────────────────────────────────────────────────────────────
# CATALOG: Excel (xlsx) from Google Drive
# Фиксированные колонки:
#   B = name, C = id, D = barcode, J = price_opt, L = price_retail
# ──────────────────────────────────────────────────────────────────────────────

def _find_catalog_start_row(sheet) -> int:
    """
    Файл может иметь “шапку” не в первой строке.
    Ищем строку, где в колонке C похоже на "Артикул", и стартуем со следующей.
    Если не нашли — стартуем со 2-й.
    """
    max_scan = min(60, sheet.max_row or 60)
    for r in range(1, max_scan + 1):
        c_val = _norm_str(sheet.cell(row=r, column=3).value)  # C
        b_val = _norm_str(sheet.cell(row=r, column=2).value)  # B
        if _is_header_like(c_val) or _is_header_like(b_val):
            # чаще всего header: C="Артикул"
            return r + 1
    return 2


def _parse_d11_catalog_excel_xlsx(data: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(io.BytesIO(data), data_only=True)
    sheet = wb.active

    start_row = _find_catalog_start_row(sheet)

    items: List[Dict[str, Any]] = []
    for r in range(start_row, (sheet.max_row or start_row) + 1):
        name = _norm_str(sheet.cell(row=r, column=2).value)     # B
        item_id = _norm_str(sheet.cell(row=r, column=3).value)  # C
        barcode = _norm_str(sheet.cell(row=r, column=4).value)  # D

        price_opt = _to_float(sheet.cell(row=r, column=10).value)     # J
        price_retail = _to_float(sheet.cell(row=r, column=12).value)  # L

        if not item_id or _is_header_like(item_id):
            continue
        if not name or _is_header_like(name):
            # бывают строки-разделы — пропускаем
            continue

        items.append(
            {
                "id": item_id,
                "name": name,
                "barcode": barcode,
                "price_opt": float(price_opt or 0.0),
                "price_retail": float(price_retail or 0.0),
            }
        )

    logger.info("D11: Каталог обработан, позиций: %d", len(items))
    return items


def _apply_catalog_delta_and_update_state(*, code: str, items: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    """
    Возвращает только изменившиеся позиции:
      [{"id": ..., "name": ..., "barcode": ...}, ...]
    И обновляет кэш:
      - sig_by_id
      - opt_by_id
      - retail_by_id
    """
    state = D11State.load(code)
    is_first_run = len(state.sig_by_id) == 0

    changed: List[Dict[str, str]] = []

    new_sig_by_id: Dict[str, str] = dict(state.sig_by_id)
    new_opt_by_id: Dict[str, float] = dict(state.opt_by_id)
    new_retail_by_id: Dict[str, float] = dict(state.retail_by_id)

    for it in items:
        item_id = _norm_str(it.get("id"))
        name = _norm_str(it.get("name"))
        barcode = _norm_str(it.get("barcode"))
        price_opt = float(it.get("price_opt") or 0.0)
        price_retail = float(it.get("price_retail") or 0.0)

        sig = _make_signature(
            item_id=item_id,
            name=name,
            barcode=barcode,
            price_opt=price_opt,
            price_retail=price_retail,
        )
        prev_sig = state.sig_by_id.get(item_id)

        if is_first_run or (prev_sig != sig):
            changed.append({"id": item_id, "name": name, "barcode": barcode})

        new_sig_by_id[item_id] = sig
        new_opt_by_id[item_id] = price_opt
        new_retail_by_id[item_id] = price_retail

    D11State(sig_by_id=new_sig_by_id, opt_by_id=new_opt_by_id, retail_by_id=new_retail_by_id).save(code)

    logger.info("D11: Delta: first_run=%s, total=%d, changed=%d", is_first_run, len(items), len(changed))
    return changed


# ──────────────────────────────────────────────────────────────────────────────
# STOCK: Excel from Google Drive (в примере .xls)
# Колонки:
#   B = code_sup (Артикул)
#   D = qty
# Цена opt/retail — из кэша каталога по code_sup (= id)
# ──────────────────────────────────────────────────────────────────────────────

def _read_xls_rows_xlrd(data: bytes) -> List[List[Any]]:
    """
    Чтение .xls через xlrd (требуется зависимость xlrd==2.0.1)
    """
    try:
        import xlrd  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "D11: Для чтения .xls нужен пакет xlrd. Установи: pip install xlrd==2.0.1"
        ) from e

    book = xlrd.open_workbook(file_contents=data)
    sh = book.sheet_by_index(0)
    rows: List[List[Any]] = []
    for r in range(sh.nrows):
        rows.append(sh.row_values(r))
    return rows


def _parse_d11_stock_excel_bytes(data: bytes, state: D11State) -> List[Dict[str, Any]]:
    rows_raw = _read_xls_rows_xlrd(data)
    out: List[Dict[str, Any]] = []

    for row in rows_raw:
        # безопасно для коротких строк
        code_sup = _norm_str(row[1] if len(row) > 1 else "")  # B
        qty_raw = row[3] if len(row) > 3 else None            # D

        if not code_sup or _is_header_like(code_sup):
            continue

        qty = _safe_int(qty_raw)
        if qty <= 0:
            continue

        price_opt = float(state.opt_by_id.get(code_sup) or 0.0)
        price_retail = float(state.retail_by_id.get(code_sup) or 0.0)

        out.append(
            {
                "code_sup": code_sup,
                "qty": qty,
                "price_retail": int(round(price_retail)) if price_retail else 0,
                "price_opt": round(price_opt, 2),
            }
        )

    logger.info("D11: Остатки обработаны, позиций qty>0: %d", len(out))
    return out


# ──────────────────────────────────────────────────────────────────────────────
# PUBLIC API
# ──────────────────────────────────────────────────────────────────────────────

async def parse_feed_catalog_to_json(*, code: str = D11_CODE_DEFAULT, timeout: int = 30) -> str:
    """
    Каталог D11:
      1) gdrive_folder из dropship_enterprises
      2) берём самый свежий файл из папки
      3) колонки B/C/D/J/L
      4) 1-й раз всё, дальше только изменения
      5) сохраняем цены opt/retail в кэш
    """
    try:
        folder_id = await _get_gdrive_folder_by_code(code)
        if not folder_id:
            raise RuntimeError(f"D11: Не найден gdrive_folder в dropship_enterprises для code='{code}'")

        drive_service = await _connect_to_google_drive()
        file_meta = await _fetch_latest_file_metadata(drive_service, folder_id)
        file_id = file_meta["id"]
        file_name = file_meta.get("name") or "catalog.xlsx"
        logger.info("D11: Каталог: найден файл: %s (%s)", file_name, file_id)

        file_bytes = await _download_file_bytes(drive_service, file_id)

        # ожидаем xlsx
        all_items = _parse_d11_catalog_excel_xlsx(file_bytes)
        changed_items = _apply_catalog_delta_and_update_state(code=code, items=all_items)

        return json.dumps(changed_items, ensure_ascii=False, indent=2)

    except Exception as e:
        msg = f"D11: Ошибка обработки каталога: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        return "[]"


async def parse_feed_stock_to_json(*, code: str = D11_CODE_DEFAULT, timeout: int = 60) -> str:
    """
    Остатки D11:
      - dropship_enterprises.feed_url содержит ID папки Google Drive
      - берём самый свежий файл из папки (обычно .xls)
      - колонки: B=Артикул(code_sup), D=qty
      - price_opt / price_retail — из кэша каталога по Артикулу
    """
    try:
        folder_id = await _get_feed_url_by_code(code)
        if not folder_id:
            raise RuntimeError(f"D11: Не найден feed_url (ID папки остатков) в dropship_enterprises для code='{code}'")

        drive_service = await _connect_to_google_drive()
        file_meta = await _fetch_latest_file_metadata(drive_service, folder_id)
        file_id = file_meta["id"]
        file_name = file_meta.get("name") or "stock.xls"
        logger.info("D11: Остатки: найден файл: %s (%s)", file_name, file_id)

        file_bytes = await _download_file_bytes(drive_service, file_id)

        state = D11State.load(code)
        stock_rows = _parse_d11_stock_excel_bytes(file_bytes, state)

        return json.dumps(stock_rows, ensure_ascii=False, indent=2)

    except Exception as e:
        msg = f"D11: Ошибка обработки остатков: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        return "[]"


async def parse_feed_to_json(
    *,
    mode: Literal["catalog", "stock"] = "catalog",
    code: str = D11_CODE_DEFAULT,
    timeout: int = 60,
) -> str:
    if mode == "catalog":
        return await parse_feed_catalog_to_json(code=code, timeout=timeout)
    if mode == "stock":
        return await parse_feed_stock_to_json(code=code, timeout=timeout)
    raise ValueError("mode must be 'catalog' or 'stock'")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description=(
            "D11 (Toros): "
            "catalog (Google Drive папка из gdrive_folder, xlsx, колонки B/C/D/J/L, full->delta, кэш цен) "
            "и stock (Google Drive папка из feed_url, xls, колонки B/D, цены из кэша каталога)."
        )
    )
    parser.add_argument("--mode", choices=["catalog", "stock"], default="catalog")
    parser.add_argument("--code", default=D11_CODE_DEFAULT)
    parser.add_argument("--timeout", type=int, default=60)

    args = parser.parse_args()
    out = asyncio.run(parse_feed_to_json(mode=args.mode, code=args.code, timeout=args.timeout))
    print(out)