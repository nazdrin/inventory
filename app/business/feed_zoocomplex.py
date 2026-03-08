from __future__ import annotations

import asyncio
import io
import json
import logging
from typing import Optional, List, Dict, Literal, Any

import httpx
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from sqlalchemy import text

from app.database import get_async_db
from app.services.notification_service import send_notification

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

ZOOCOMPLEX_CODE_DEFAULT = "D13"
ZOOCOMPLEX_FEED_URL = "https://zoocomplex.com.ua/drop-feed-14qr3m3z.xml"
ZOOCOMPLEX_SHEET_ID = "1b6rodQcrqmSORuG4AdlGbYYkycHO9gWx0_X0Pl73oEQ"
ZOOCOMPLEX_SHEET_GID = "2101479018"


def _get_text(el: ET.Element, candidates: List[str]) -> Optional[str]:
    for tag in candidates:
        child = el.find(tag)
        if child is not None and child.text and child.text.strip():
            return child.text.strip()
    return None


def _to_float(val: Optional[Any]) -> float:
    if val is None:
        return 0.0
    s = str(val).strip().replace("\u00A0", "").replace(" ", "").replace(",", ".")
    if not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def _normalize_barcode(value: Optional[Any]) -> str:
    if value is None:
        return ""
    barcode = str(value).strip().replace("\u00A0", "").replace(" ", "")
    return barcode


async def _get_feed_url_by_code(code: str) -> Optional[str]:
    async with get_async_db() as session:
        res = await session.execute(
            text("SELECT feed_url FROM dropship_enterprises WHERE code = :code LIMIT 1"),
            {"code": code},
        )
        feed_url = res.scalar_one_or_none()
    if feed_url:
        return str(feed_url).strip()
    return ZOOCOMPLEX_FEED_URL


async def _get_profit_percent_by_code(code: str) -> float:
    async with get_async_db() as session:
        res = await session.execute(
            text("SELECT profit_percent FROM dropship_enterprises WHERE code = :code LIMIT 1"),
            {"code": code},
        )
        raw = res.scalar_one_or_none()

    try:
        val = float(raw)
    except Exception:
        return 0.0

    if val > 1:
        val = val / 100.0
    if val < 0:
        val = 0.0
    if val > 1:
        val = 1.0
    return val


async def _load_feed_root(*, code: str, timeout: int) -> Optional[ET.Element]:
    feed_url = await _get_feed_url_by_code(code)
    if not feed_url:
        msg = f"Не найден feed_url для code='{code}'"
        logger.error(msg)
        send_notification(msg, "Разработчик")
        return None

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "application/xml,text/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "uk-UA,uk;q=0.9,en-US;q=0.8,en;q=0.7",
        "Referer": "https://zoocomplex.com.ua/",
        "Connection": "keep-alive",
    }
    candidate_urls: List[str] = []
    base = str(feed_url).strip()
    if base:
        candidate_urls.append(base)
        if "://zoocomplex.com.ua/" in base:
            candidate_urls.append(base.replace("://zoocomplex.com.ua/", "://www.zoocomplex.com.ua/"))
            candidate_urls.append(base.replace("https://", "http://"))
        if "://www.zoocomplex.com.ua/" in base:
            candidate_urls.append(base.replace("://www.zoocomplex.com.ua/", "://zoocomplex.com.ua/"))
            candidate_urls.append(base.replace("https://", "http://"))
    # unique while preserving order
    candidate_urls = list(dict.fromkeys(candidate_urls))

    last_error: Optional[Exception] = None
    xml_text: Optional[str] = None
    try:
        async with httpx.AsyncClient(headers=headers, timeout=timeout, follow_redirects=True) as client:
            for url in candidate_urls:
                try:
                    resp = await client.get(url)
                    if resp.status_code == 403:
                        logger.warning("Zoocomplex feed returned 403 for url=%s", url)
                        continue
                    resp.raise_for_status()
                    xml_text = resp.text
                    if xml_text and "<offer" in xml_text:
                        logger.info("Zoocomplex feed loaded from %s", url)
                        break
                    logger.warning("Zoocomplex feed loaded but has no offer nodes, url=%s", url)
                except Exception as e:
                    last_error = e
                    logger.warning("Zoocomplex feed attempt failed for %s: %s", url, e)
                    continue
    except Exception as e:
        last_error = e

    if not xml_text:
        err_text = str(last_error) if last_error else "unknown error"
        msg = f"Ошибка загрузки фида Zoocomplex {feed_url}: {err_text}"
        logger.error(msg)
        send_notification(msg, "Разработчик")
        return None

    try:
        root = ET.fromstring(xml_text)
    except Exception as e:
        msg = f"Ошибка парсинга XML Zoocomplex {feed_url}: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        return None

    offers = root.findall(".//offer")
    if not offers:
        msg = f"Zoocomplex feed пустой или не содержит offer (url={feed_url})"
        logger.error(msg)
        send_notification(msg, "Разработчик")
        return None

    return root


def _collect_offer_nodes(root: ET.Element) -> List[ET.Element]:
    offers = root.findall(".//offer")
    if not offers:
        offers = [el for el in root.iter() if el.tag.lower() == "offer"]
    return offers


def _extract_offer_id(offer: ET.Element) -> Optional[str]:
    value = offer.get("id") or _get_text(offer, ["id"])
    if not value:
        return None
    return str(value).strip()


def _extract_name(offer: ET.Element) -> Optional[str]:
    value = _get_text(offer, ["name", "name_ua", "title"])
    if not value:
        return None
    return " ".join(value.split())


def _extract_barcode(offer: ET.Element) -> str:
    return _normalize_barcode(_get_text(offer, ["barcode", "ean", "gtin", "upc"]))


def _extract_available_qty(offer: ET.Element) -> int:
    available = str(offer.get("available") or "").strip().lower()
    return 1 if available == "true" else 0


def _extract_price_retail(offer: ET.Element) -> float:
    return _to_float(_get_text(offer, ["price"]))


async def _download_google_sheet_xlsx(*, timeout: int = 30) -> bytes:
    url = (
        f"https://docs.google.com/spreadsheets/d/{ZOOCOMPLEX_SHEET_ID}/export"
        f"?format=xlsx&gid={ZOOCOMPLEX_SHEET_GID}"
    )
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        async with httpx.AsyncClient(
            headers=headers,
            timeout=timeout,
            follow_redirects=True,
        ) as client:
            resp = await client.get(url)
            resp.raise_for_status()
            return resp.content
    except Exception as e:
        msg = f"Ошибка загрузки Google Sheet XLSX для Zoocomplex: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")
        raise


def _load_wholesale_price_map(xlsx_bytes: bytes) -> Dict[str, float]:
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
        sheet = wb.active
    except Exception as e:
        raise ValueError(f"Не удалось открыть XLSX Zoocomplex: {e}")

    mapping: Dict[str, float] = {}
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row:
            continue

        barcode_val = row[3] if len(row) > 3 else None  # D
        price_val = row[6] if len(row) > 6 else None    # G

        barcode = _normalize_barcode(barcode_val)
        if not barcode:
            continue

        price_opt = _to_float(price_val)
        if price_opt <= 0:
            continue

        mapping[barcode] = price_opt

    logger.info("Zoocomplex wholesale map loaded: %d barcodes", len(mapping))
    return mapping


async def parse_zoocomplex_catalog_to_json(*, code: str = ZOOCOMPLEX_CODE_DEFAULT, timeout: int = 30) -> str:
    root = await _load_feed_root(code=code, timeout=timeout)
    if root is None:
        return "[]"

    items: List[Dict[str, str]] = []
    for offer in _collect_offer_nodes(root):
        offer_id = _extract_offer_id(offer)
        name = _extract_name(offer)
        barcode = _extract_barcode(offer)

        if not (offer_id and name):
            continue

        items.append({
            "id": offer_id,
            "name": name,
            "barcode": barcode,
        })

    logger.info("Zoocomplex catalog parsed (code=%s): %d", code, len(items))
    return json.dumps(items, ensure_ascii=False, indent=2)


async def parse_zoocomplex_stock_to_json(*, code: str = ZOOCOMPLEX_CODE_DEFAULT, timeout: int = 30) -> str:
    root = await _load_feed_root(code=code, timeout=timeout)
    if root is None:
        return "[]"

    profit_percent = await _get_profit_percent_by_code(code)

    wholesale_map: Dict[str, float] = {}
    try:
        xlsx_bytes = await _download_google_sheet_xlsx(timeout=timeout)
        wholesale_map = _load_wholesale_price_map(xlsx_bytes)
    except Exception as e:
        msg = f"Zoocomplex wholesale map unavailable, fallback to profit formula: {e}"
        logger.exception(msg)
        send_notification(msg, "Разработчик")

    rows: List[Dict[str, Any]] = []
    for offer in _collect_offer_nodes(root):
        offer_id = _extract_offer_id(offer)
        if not offer_id:
            continue

        qty = _extract_available_qty(offer)
        price_retail = _extract_price_retail(offer)
        barcode = _extract_barcode(offer)

        price_opt = wholesale_map.get(barcode, 0.0) if barcode else 0.0
        if price_opt <= 0.0:
            price_opt = price_retail / (1.0 + profit_percent) if price_retail > 0 else 0.0
        if price_opt < 0:
            price_opt = 0.0

        rows.append(
            {
                "code_sup": offer_id,
                "qty": qty,
                "price_retail": price_retail,
                "price_opt": price_opt,
            }
        )

    logger.info("Zoocomplex stock parsed (code=%s): %d", code, len(rows))
    return json.dumps(rows, ensure_ascii=False, indent=2)


async def parse_zoocomplex_feed_to_json(
    *,
    mode: Literal["catalog", "stock"] = "catalog",
    code: str = ZOOCOMPLEX_CODE_DEFAULT,
    timeout: int = 30,
) -> str:
    if mode == "catalog":
        return await parse_zoocomplex_catalog_to_json(code=code, timeout=timeout)
    if mode == "stock":
        return await parse_zoocomplex_stock_to_json(code=code, timeout=timeout)
    raise ValueError("mode must be 'catalog' or 'stock'")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Zoocomplex parser")
    parser.add_argument("--mode", choices=["catalog", "stock"], default="catalog")
    parser.add_argument("--code", default=ZOOCOMPLEX_CODE_DEFAULT)
    parser.add_argument("--timeout", type=int, default=30)
    args = parser.parse_args()

    result = asyncio.run(
        parse_zoocomplex_feed_to_json(mode=args.mode, code=args.code, timeout=args.timeout)
    )
    print(result)
